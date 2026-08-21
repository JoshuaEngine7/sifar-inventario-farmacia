from datetime import date, datetime, time, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from . import config, services
from .auth import require_user
from .plantillas import templates
from .db import get_session
from .models import Producto, Usuario

router = APIRouter()


def _rango(desde: str, hasta: str) -> tuple[date, date]:
    hoy = date.today()
    try:
        d = datetime.strptime(desde, "%Y-%m-%d").date() if desde else hoy.replace(day=1)
        h = datetime.strptime(hasta, "%Y-%m-%d").date() if hasta else hoy
    except ValueError:
        return hoy.replace(day=1), hoy
    return (d, h) if d <= h else (h, d)


def _filas_reporte(session: Session, d: date, h: date, solo_movimientos: bool):
    inicio = datetime.combine(d, time.min)
    fin = datetime.combine(h, time.max)
    stock_inicial = services.stock_en_fecha_todos(session, inicio - timedelta(seconds=1))
    stock_final = services.stock_en_fecha_todos(session, fin)
    sumas = services.sumas_periodo(session, inicio, fin)

    productos = (
        session.query(Producto)
        .options(joinedload(Producto.tipo), joinedload(Producto.ubicacion))
        .filter(Producto.activo.is_(True))
        .order_by(Producto.nombre)
        .all()
    )
    filas = []
    for p in productos:
        s = sumas.get(p.id, {"entradas": 0, "salidas": 0, "ajustes": 0})
        if solo_movimientos and not any(s.values()):
            continue
        filas.append({
            "p": p,
            "inicial": stock_inicial.get(p.id, 0),
            "final": stock_final.get(p.id, 0),
            **s,
        })
    return filas


@router.get("/reporte")
def reporte(request: Request, desde: str = "", hasta: str = "", solo_mov: int = 0,
            session: Session = Depends(get_session),
            usuario: Usuario = Depends(require_user)):
    d, h = _rango(desde, hasta)
    filas = _filas_reporte(session, d, h, bool(solo_mov))
    return templates.TemplateResponse(
        request, "reporte.html",
        {"usuario": usuario, "filas": filas, "desde": d, "hasta": h,
         "solo_mov": bool(solo_mov)},
    )


@router.get("/reporte/export")
def exportar(desde: str = "", hasta: str = "", solo_mov: int = 0,
             session: Session = Depends(get_session),
             usuario: Usuario = Depends(require_user)):
    d, h = _rango(desde, hasta)
    filas = _filas_reporte(session, d, h, bool(solo_mov))

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws["A1"] = config.TITULO_REPORTE.upper()
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Periodo: {d:%d/%m/%Y} al {h:%d/%m/%Y}"
    encabezados = ["Producto", "Tipo", "Ubicación", "Stock inicial",
                   "Entradas", "Salidas", "Ajustes", "Stock final"]
    ws.append([])
    ws.append(encabezados)
    for celda in ws[4]:
        celda.font = Font(bold=True)
    # El export no incluye pacientes ni causas: es agregado por producto.
    # texto_excel: los nombres de producto ahora los crean usuarios (inyección de fórmulas).
    for i, f in enumerate(filas, start=5):
        ws.append([services.texto_excel(f["p"].nombre), f["p"].tipo.nombre,
                   f["p"].ubicacion.nombre,
                   f["inicial"], f["entradas"], f["salidas"], f["ajustes"], None])
        ws.cell(row=i, column=8).value = f"=D{i}+E{i}-F{i}+G{i}"
    for col, ancho in zip("ABCDEFGH", (52, 22, 18, 12, 10, 10, 10, 11)):
        ws.column_dimensions[col].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"reporte_{d:%Y%m%d}_{h:%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
