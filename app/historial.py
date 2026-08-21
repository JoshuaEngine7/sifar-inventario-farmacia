"""Bitácora de movimientos: quién registró qué, cuándo, con causa y paciente.

Es la vista de CONTROL INTERNO (visible para todo usuario con sesión, decisión
del doctor en el cuestionario). Complementa al Reporte por periodo, cuyo Excel
va SIN pacientes porque se entrega a certificaciones.
"""
from datetime import date, datetime, time
from io import BytesIO

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session, joinedload

from . import config, services
from .auth import require_user
from .db import get_session
from .models import TIPOS_MOVIMIENTO, Movimiento, Producto, Usuario
from .plantillas import templates

router = APIRouter()

LIMITE_FILAS = 300        # en pantalla
LIMITE_EXPORT = 10000     # en el Excel


def _fecha(valor: str) -> date | None:
    try:
        return datetime.strptime(valor.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _consulta(session: Session, q: str, producto_id: str, tipo: str,
              desde: str, hasta: str):
    consulta = (
        session.query(Movimiento)
        .join(Movimiento.producto)
        .options(joinedload(Movimiento.producto).joinedload(Producto.ubicacion),
                 joinedload(Movimiento.causa), joinedload(Movimiento.usuario))
    )
    # Campos como str + parseo manual: los formularios mandan "" (lección aprendida).
    # isdecimal, no isdigit: '²'.isdigit() es True pero int('²') truena con 500.
    if producto_id.isdecimal():
        consulta = consulta.filter(Movimiento.producto_id == int(producto_id))
    if q.strip():
        consulta = consulta.filter(Producto.nombre.ilike(f"%{q.strip()}%"))
    if tipo in TIPOS_MOVIMIENTO:
        consulta = consulta.filter(Movimiento.tipo == tipo)
    d, h = _fecha(desde), _fecha(hasta)
    if d:
        consulta = consulta.filter(Movimiento.fecha_hora >= datetime.combine(d, time.min))
    if h:
        consulta = consulta.filter(Movimiento.fecha_hora <= datetime.combine(h, time.max))
    return consulta.order_by(Movimiento.fecha_hora.desc(), Movimiento.id.desc())


@router.get("/historial")
def historial(request: Request, q: str = "", producto_id: str = "", tipo: str = "",
              desde: str = "", hasta: str = "",
              session: Session = Depends(get_session),
              usuario: Usuario = Depends(require_user)):
    consulta = _consulta(session, q, producto_id, tipo, desde, hasta)
    total = consulta.count()
    filas = consulta.limit(LIMITE_FILAS).all()
    producto_fijado = (session.get(Producto, int(producto_id))
                       if producto_id.isdecimal() else None)
    return templates.TemplateResponse(
        request, "historial.html",
        {
            "usuario": usuario, "filas": filas, "total": total,
            "limite": LIMITE_FILAS, "tipos": TIPOS_MOVIMIENTO,
            "q": q, "producto_id": producto_id, "tipo": tipo,
            "desde": desde, "hasta": hasta,
            "producto_fijado": producto_fijado,
        },
    )


@router.get("/historial/export")
def exportar(q: str = "", producto_id: str = "", tipo: str = "",
             desde: str = "", hasta: str = "",
             session: Session = Depends(get_session),
             usuario: Usuario = Depends(require_user)):
    consulta = _consulta(session, q, producto_id, tipo, desde, hasta)
    total = consulta.count()
    filas = consulta.limit(LIMITE_EXPORT).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"
    ws["A1"] = f"{config.APP_NOMBRE} — {config.UNIDAD} — BITÁCORA DE MOVIMIENTOS".upper()
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "USO INTERNO — contiene datos de pacientes; no entregar fuera del consultorio"
    ws["A2"].font = Font(bold=True, color="A8322A")
    if total > LIMITE_EXPORT:
        # Nunca truncar en silencio: una bitácora incompleta que se cree completa
        # es peor que no tenerla.
        ws["A3"] = (f"AVISO: se exportaron {LIMITE_EXPORT:,} de {total:,} movimientos "
                    f"(los más recientes). Acote el rango de fechas para ver el resto.")
        ws["A3"].font = Font(bold=True, color="A8322A")
    else:
        ws.append([])
    encabezados = ["Fecha y hora", "Producto", "Ubicación", "Tipo", "Piezas", "Cajas",
                   "Caducidad", "Causa", "Detalle", "Paciente", "Registró", "Origen"]
    ws.append(encabezados)
    for celda in ws[4]:
        celda.font = Font(bold=True)
    txt = services.texto_excel
    for m in filas:
        ws.append([
            m.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            txt(m.producto.nombre),
            txt(m.producto.ubicacion.nombre),
            m.tipo,
            m.piezas,
            m.cajas,
            m.fecha_caducidad.strftime("%d/%m/%Y") if m.fecha_caducidad else "",
            txt(m.causa.nombre if m.causa else ""),
            txt(m.causa_detalle),
            txt(m.paciente_ref),
            txt(m.usuario.nombre),
            "Excel migrado" if m.historico else "SIFAR",
        ])
    for col, ancho in zip("ABCDEFGHIJKL", (16, 45, 16, 14, 8, 7, 11, 24, 26, 28, 26, 13)):
        ws.column_dimensions[col].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"historial_{date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
