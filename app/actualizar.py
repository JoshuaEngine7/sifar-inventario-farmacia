"""Actualización INCREMENTAL desde un Excel con los stocks actuales.

A diferencia de la migración (que solo corre sobre una base vacía), esto corre
sobre una base EN USO y nunca borra nada:

  - Producto del Excel que no existe en SIFAR  → se da de alta (con sus lotes).
  - Stock de piezas distinto                   → movimiento AJUSTE auditado.
  - Cajas de un lote distintas                 → movimiento AJUSTE auditado.
  - Lo que SIFAR tiene y el Excel no menciona  → SE CONSERVA intacto (se reporta).
  - Movimientos, usuarios y contraseñas        → nunca se tocan.

Uso:
  python -m app.actualizar maybe.xlsm            (vista previa: no cambia nada)
  python -m app.actualizar maybe.xlsm --aplicar  (aplica; respalda antes)

Genera actualizacion_reporte.txt con el detalle.
"""
import sys
from datetime import date

from openpyxl import load_workbook

from . import respaldo, services
from .db import SessionLocal
from .migrate import MAPA_TIPO_UBIC, TIPOS_MATERIAL, _fecha, _int, parse_piezas_por_caja
from .models import Causa, Lote, Producto, Tipo, Ubicacion, Usuario


def analizar(session, ws) -> dict:
    tipos = {t.nombre: t for t in session.query(Tipo)}
    ubicaciones = {u.nombre: u for u in session.query(Ubicacion)}
    indice = {(p.tipo_id, p.ubicacion_id, p.nombre): p for p in session.query(Producto)}
    stocks = services.stock_piezas_todos(session)

    plan = {"nuevos": [], "ajustes_piezas": [], "ajustes_cajas": [],
            "sin_cambio": 0, "avisos": []}
    vistos = set()
    filas_procesadas = set()

    for n_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tipo_ubic = str(fila[0] or "").strip()
        nombre = str(fila[1] or "").strip()
        if not nombre:
            continue
        if tipo_ubic not in MAPA_TIPO_UBIC:
            plan["avisos"].append(f"fila {n_fila}: Tipo-Ubicación desconocido '{tipo_ubic}' — omitida")
            continue
        if (tipo_ubic, nombre) in filas_procesadas:
            # El Excel real trae renglones duplicados; procesar el segundo
            # aplicaría el mismo ajuste dos veces.
            plan["avisos"].append(f"fila {n_fila}: '{nombre}' duplicado en el Excel — se usa solo el primero")
            continue
        filas_procesadas.add((tipo_ubic, nombre))
        if fila[5] is None:
            # Columna Stock Final vacía: suele ser un Excel guardado sin calcular
            # fórmulas. Ajustar con 0 sería destructivo; mejor omitir y avisar.
            plan["avisos"].append(f"fila {n_fila}: '{nombre}' sin valor en Stock Final — omitida")
            continue

        nombre_tipo, nombre_ubic = MAPA_TIPO_UBIC[tipo_ubic]
        tipo, ubicacion = tipos[nombre_tipo], ubicaciones[nombre_ubic]
        stock_excel = _int(fila[5], plan["avisos"], f"fila {n_fila} Stock Final ({nombre})")
        lotes_excel: dict[date, int] = {}
        for col in range(6, 16, 2):
            fecha = _fecha(fila[col + 1])
            if fecha is not None:
                cajas = _int(fila[col], plan["avisos"], f"fila {n_fila} col {col + 1} ({nombre})")
                lotes_excel[fecha] = lotes_excel.get(fecha, 0) + cajas

        producto = indice.get((tipo.id, ubicacion.id, nombre))
        if producto is None:
            plan["nuevos"].append({
                "nombre": nombre, "tipo": tipo, "ubicacion": ubicacion,
                "stock": stock_excel, "lotes": lotes_excel,
                "unidad": "caja" if nombre_tipo in TIPOS_MATERIAL and stock_excel == 0 else "pieza",
                "ppc": parse_piezas_por_caja(nombre),
            })
            continue

        vistos.add(producto.id)
        cambio = False

        delta = stock_excel - stocks.get(producto.id, 0)
        if delta != 0:
            if producto.unidad == "caja":
                plan["avisos"].append(
                    f"'{nombre}': controla por caja; diferencia de piezas ({delta}) se ignora")
            else:
                plan["ajustes_piezas"].append(
                    {"producto": producto, "delta": delta,
                     "excel": stock_excel, "actual": stocks.get(producto.id, 0)})
                cambio = True

        lotes_sifar = {l.fecha_caducidad: l for l in producto.lotes}
        for fecha, cajas_excel in lotes_excel.items():
            lote = lotes_sifar.get(fecha)
            delta_cajas = cajas_excel - (lote.cajas if lote else 0)
            if delta_cajas != 0:
                plan["ajustes_cajas"].append(
                    {"producto": producto, "fecha": fecha, "delta": delta_cajas,
                     "lote": lote, "excel": cajas_excel})
                cambio = True
        for fecha in set(lotes_sifar) - set(lotes_excel):
            if lotes_sifar[fecha].cajas > 0:
                plan["avisos"].append(
                    f"'{nombre}': lote {fecha} ({lotes_sifar[fecha].cajas} cajas) no está "
                    f"en el Excel — SE CONSERVA tal cual")
        if not cambio:
            plan["sin_cambio"] += 1

    plan["no_en_excel"] = sum(1 for p in indice.values()
                              if p.id not in vistos and p.activo)
    return plan


def aplicar(session, plan) -> list[str]:
    hoy = date.today()
    detalle_txt = f"Actualización desde Excel {hoy:%d/%m/%Y}"
    usuario = session.query(Usuario).filter_by(nombre="Migración Excel").one()
    causa = session.query(Causa).filter_by(nombre="Ajuste por conteo").one()
    errores = []

    for n in plan["nuevos"]:
        producto = Producto(nombre=n["nombre"], tipo_id=n["tipo"].id,
                            ubicacion_id=n["ubicacion"].id, unidad=n["unidad"],
                            piezas_por_caja=n["ppc"], stock_base=n["stock"])
        session.add(producto)
        for fecha, cajas in n["lotes"].items():
            session.add(Lote(producto=producto, cajas=cajas, fecha_caducidad=fecha))

    # Cada ajuste va en un SAVEPOINT: si el servicio lo rechaza, se revierte
    # SOLO ese item (el servicio muta y valida después; sin esto, un rechazo
    # dejaría su lote/movimiento a medias dentro de la transacción grande).
    for a in plan["ajustes_piezas"]:
        try:
            with session.begin_nested():
                services.registrar_movimiento(
                    session, usuario=usuario, producto=a["producto"], tipo="AJUSTE",
                    piezas=a["delta"], causa=causa, causa_detalle=detalle_txt)
        except services.MovimientoInvalido as exc:
            errores.append(f"piezas '{a['producto'].nombre}': {exc}")

    for a in plan["ajustes_cajas"]:
        try:
            with session.begin_nested():
                lote = a["lote"]
                if lote is None:
                    lote = Lote(producto_id=a["producto"].id, cajas=0,
                                fecha_caducidad=a["fecha"])
                    session.add(lote)
                    session.flush()
                services.registrar_movimiento(
                    session, usuario=usuario, producto=a["producto"], tipo="AJUSTE",
                    cajas=a["delta"], lote=lote, causa=causa, causa_detalle=detalle_txt)
        except services.MovimientoInvalido as exc:
            errores.append(f"cajas '{a['producto'].nombre}' lote {a['fecha']}: {exc}")

    session.commit()
    return errores


def _resumen(plan) -> str:
    lineas = [
        f"Productos NUEVOS a dar de alta: {len(plan['nuevos'])}",
        *(f"   + {n['nombre']} ({n['tipo'].nombre} / {n['ubicacion'].nombre}) "
          f"stock {n['stock']}, lotes {len(n['lotes'])}" for n in plan["nuevos"]),
        f"AJUSTES de piezas: {len(plan['ajustes_piezas'])}",
        *(f"   ~ {a['producto'].nombre}: {a['actual']} -> {a['excel']} ({a['delta']:+d})"
          for a in plan["ajustes_piezas"]),
        f"AJUSTES de cajas por lote: {len(plan['ajustes_cajas'])}",
        *(f"   ~ {a['producto'].nombre} lote {a['fecha']}: {a['delta']:+d} cajas"
          for a in plan["ajustes_cajas"]),
        f"Sin cambios: {plan['sin_cambio']} productos",
        f"En SIFAR pero no en el Excel (se conservan intactos): {plan['no_en_excel']}",
        f"Avisos: {len(plan['avisos'])}",
        *(f"   ! {a}" for a in plan["avisos"]),
    ]
    return "\n".join(lineas)


def run(ruta_excel: str, aplicar_cambios: bool) -> None:
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    with SessionLocal() as session:
        plan = analizar(session, wb["INVENTARIO"])
        resumen = _resumen(plan)
        print(resumen)

        if not aplicar_cambios:
            print("\n(VISTA PREVIA: no se cambió nada. Para aplicar: --aplicar)")
        else:
            copia = respaldo.run("respaldo")
            print(f"\nRespaldo previo creado: {copia}")
            errores = aplicar(session, plan)
            print("Cambios APLICADOS." if not errores else
                  f"Aplicado con {len(errores)} rechazos:\n" + "\n".join(f"   ! {e}" for e in errores))
            resumen += "\n\nAPLICADO: sí\nRechazos:\n" + ("\n".join(errores) or "ninguno")

    with open("actualizacion_reporte.txt", "w", encoding="utf-8") as f:
        f.write(resumen)
    print("Detalle en actualizacion_reporte.txt")


if __name__ == "__main__":
    argumentos = [a for a in sys.argv[1:] if a != "--aplicar"]
    if len(argumentos) != 1:
        sys.exit('Uso: python -m app.actualizar "maybe.xlsm" [--aplicar]')
    run(argumentos[0], "--aplicar" in sys.argv)
