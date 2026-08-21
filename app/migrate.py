"""Migración desde maybe.xlsm: productos + lotes (INVENTARIO) e historial (HISTORIAL_).

Uso:  python -m app.migrate "excel original/maybe.xlsm"

Corre sobre una BD recién sembrada (python -m app.seed). Si ya hay productos,
aborta para no duplicar: borrar farmacia.db y re-sembrar para repetir.
Genera migracion_revision.txt con todo lo que requiere ojo humano.
"""
import re
import sys
from datetime import date, datetime

from openpyxl import load_workbook

from .db import SessionLocal
from .models import Causa, Lote, Movimiento, Producto, Tipo, Ubicacion, Usuario
from . import seed

# "Tipo - Ubicación" del Excel → (tipo, ubicación) confirmados por el doctor (Q7-8).
# Carro rojo: tipo por defecto Medicamento (editable después por admin).
MAPA_TIPO_UBIC = {
    "MEDICAMENTO EN FARMACIA": ("Medicamento", "Farmacia"),
    "ANTIBIOTICO EN FARMACIA": ("Antibiótico", "Farmacia"),
    "ANTIBIOTICO INYECTABLE EN FARMACIA": ("Antibiótico inyectable", "Farmacia"),
    "INYECTABLE EN FARMACIA": ("Inyectable", "Farmacia"),
    "MEDICAMENTO CONTROLADO": ("Medicamento controlado", "Farmacia controlado"),
    "MATERIAL": ("Material", "Farmacia"),
    "MATERIAL RPBI": ("Material RPBI", "Farmacia"),
    "CONSULTORIO DENTAL": ("Material dental", "Consultorio dental"),
    "CARRO ROJO P1": ("Medicamento", "Carro rojo P1"),
    "CARRO ROJO P2": ("Medicamento", "Carro rojo P2"),
    "CARRO ROJO P3": ("Medicamento", "Carro rojo P3"),
    "CARRO ROJO P4": ("Medicamento", "Carro rojo P4"),
    "CARRO ROJO P5": ("Medicamento", "Carro rojo P5"),
}
TIPOS_MATERIAL = {"Material", "Material RPBI", "Material dental"}

RE_PIEZAS_CAJA = re.compile(
    r"(?:C/\s*(\d+)|CAJA CON (\d+)|FRASCO CON (\d+)"
    r"|CON (\d+)\s*(?:TAB|CAP|AMP|PIEZ|PZ|COMPRIMID|SOBRE|GRAGEA)"
    r"|(\d+)\s*PZS|(\d+)\s*PIEZAS)",
    re.IGNORECASE,
)


def parse_piezas_por_caja(nombre: str) -> int | None:
    m = RE_PIEZAS_CAJA.search(nombre)
    if not m:
        return None
    valor = int(next(g for g in m.groups() if g))
    return valor if valor > 0 else None


def parse_causa(texto: str | None) -> tuple[str | None, str | None, str | None]:
    """→ (nombre_causa, paciente_ref, causa_detalle)."""
    if not texto or not str(texto).strip():
        return None, None, None
    t = str(texto).strip()
    mayus = t.upper()
    m = re.match(r"^(?:TX\.?:?|TRATAMIENTO)\s*(.*)$", t, re.IGNORECASE)
    if m:
        paciente = m.group(1).strip() or None
        return "Tratamiento / receta a paciente", paciente, None
    if "CADUC" in mayus:
        return "Caducidad", None, t
    if "DONA" in mayus:
        return "Donación", None, t
    if "DEVOL" in mayus:
        return "Devolución", None, t
    if "AJUSTE" in mayus or "CONTEO" in mayus:
        return "Ajuste por conteo", None, t
    return None, None, t  # sin clasificar: se conserva íntegro en causa_detalle


def _int(valor, avisos: list[str] | None = None, contexto: str = "") -> int:
    if valor is None or valor == "":
        return 0
    try:
        return int(round(float(valor)))
    except (TypeError, ValueError):
        # Celda con texto donde debía haber número: avisar, no perder en silencio.
        if avisos is not None:
            avisos.append(f"VALOR NO NUMÉRICO tratado como 0: '{valor}' en {contexto}")
        return 0


def _fecha(valor) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def migrar_productos(session, ws, avisos: list[str]) -> dict[tuple[str, str], Producto]:
    tipos = {t.nombre: t for t in session.query(Tipo)}
    ubicaciones = {u.nombre: u for u in session.query(Ubicacion)}
    indice: dict[tuple[str, str], Producto] = {}

    # Sin max_row fijo: el Excel de la clínica seguirá creciendo hasta el día de
    # la migración real; un límite fijo truncaría filas nuevas EN SILENCIO.
    for n_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tipo_ubic = str(fila[0] or "").strip()
        nombre = str(fila[1] or "").strip()
        if not nombre:
            continue
        if tipo_ubic not in MAPA_TIPO_UBIC:
            avisos.append(f"TIPO-UBICACION desconocido: '{tipo_ubic}' ({nombre}) — se omite")
            continue
        clave_idx = (tipo_ubic, nombre)
        if clave_idx in indice:
            avisos.append(f"PRODUCTO DUPLICADO en Excel: {tipo_ubic} | {nombre} — se conserva el primero")
            continue

        nombre_tipo, nombre_ubic = MAPA_TIPO_UBIC[tipo_ubic]
        stock_final = _int(fila[5], avisos, f"INVENTARIO fila {n_fila} Stock Final ({nombre})")
        # Material/insumos se controlan por caja (Q13), salvo los que el propio
        # Excel ya contaba por pieza (tienen stock de piezas distinto de 0).
        if nombre_tipo in TIPOS_MATERIAL and stock_final == 0:
            unidad = "caja"
        else:
            unidad = "pieza"

        producto = Producto(
            nombre=nombre,
            tipo_id=tipos[nombre_tipo].id,
            ubicacion_id=ubicaciones[nombre_ubic].id,
            unidad=unidad,
            piezas_por_caja=parse_piezas_por_caja(nombre),
            stock_base=stock_final,
        )
        session.add(producto)
        indice[clave_idx] = producto

        # Lotes: pares (Cajas i, Fecha i) en columnas G..P (índices 6..15).
        lotes: dict[date, int] = {}
        for col in range(6, 16, 2):
            cajas = _int(fila[col], avisos, f"INVENTARIO fila {n_fila} col {col + 1} ({nombre})")
            fecha = _fecha(fila[col + 1])
            if fecha is None:
                continue
            if cajas < 0:
                avisos.append(f"LOTE con cajas negativas en Excel: {nombre} {fecha} ({cajas}) — se importa tal cual")
            lotes[fecha] = lotes.get(fecha, 0) + cajas
        for fecha, cajas in lotes.items():
            session.add(Lote(producto=producto, cajas=cajas, fecha_caducidad=fecha))

    session.flush()
    return indice


def migrar_historial(session, ws, indice, avisos: list[str]) -> int:
    causas = {c.nombre: c for c in session.query(Causa)}
    usuario_migracion = session.query(Usuario).filter_by(nombre="Migración Excel").one()
    tipos = {t.nombre: t for t in session.query(Tipo)}
    ubicaciones = {u.nombre: u for u in session.query(Ubicacion)}
    importados = 0

    # Sin max_row fijo (mismo motivo que en productos). min_row=2: la fila 2 va
    # vacía en el archivo actual y el filtro de nombre la salta, pero un archivo
    # futuro podría traer datos ahí.
    for n_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fecha, lugar, nombre, ent, sal, caj, fcad, causa_txt = fila[:8]
        nombre = str(nombre or "").strip()
        lugar = str(lugar or "").strip()
        if not nombre:
            continue
        if not isinstance(fecha, (datetime, date)):
            # Sin fecha no se puede ubicar el movimiento en el tiempo; inventarlo
            # (p.ej. "hoy") contaminaría reportes. Se avisa para captura manual.
            avisos.append(f"HISTORIAL fila {n_fila} SIN FECHA VÁLIDA — no importada, "
                          f"capturar a mano: {lugar} | {nombre} | causa: {causa_txt}")
            continue

        producto = indice.get((lugar, nombre))
        if producto is None:
            # Producto del historial que ya no está en INVENTARIO: se crea inactivo
            # para no perder el movimiento (el historial nunca se descarta).
            if lugar in MAPA_TIPO_UBIC:
                nombre_tipo, nombre_ubic = MAPA_TIPO_UBIC[lugar]
            else:
                nombre_tipo, nombre_ubic = "Medicamento", "Farmacia"
            producto = Producto(
                nombre=nombre, tipo_id=tipos[nombre_tipo].id,
                ubicacion_id=ubicaciones[nombre_ubic].id, activo=False,
            )
            session.add(producto)
            session.flush()
            indice[(lugar, nombre)] = producto
            avisos.append(f"HISTORIAL fila {n_fila}: producto no está en INVENTARIO, creado inactivo: {lugar} | {nombre}")

        ctx = f"HISTORIAL fila {n_fila} ({nombre})"
        entrada, salida, cajas = _int(ent, avisos, ctx), _int(sal, avisos, ctx), _int(caj, avisos, ctx)
        tipo_mov = "ENTRADA" if (entrada > 0 or (entrada == 0 and salida == 0 and cajas > 0)) else "SALIDA"
        nombre_causa, paciente, detalle = parse_causa(causa_txt)

        session.add(Movimiento(
            fecha_hora=fecha if isinstance(fecha, datetime) else datetime.combine(fecha, datetime.min.time()),
            producto_id=producto.id,
            tipo=tipo_mov,
            piezas=entrada if tipo_mov == "ENTRADA" else salida,
            cajas=cajas,  # con signo, tal como el Excel (negativo = caja que salió/se vació)
            fecha_caducidad=_fecha(fcad),
            causa_id=causas[nombre_causa].id if nombre_causa else None,
            causa_detalle=detalle,
            paciente_ref=paciente,
            usuario_id=usuario_migracion.id,
            historico=True,
        ))
        importados += 1

    return importados


def run(ruta_excel: str) -> None:
    seed.run()  # garantiza tablas y catálogos
    with SessionLocal() as session:
        if session.query(Producto).count() > 0:
            sys.exit("La BD ya tiene productos. Borra farmacia.db y vuelve a correr seed + migrate.")

        wb = load_workbook(ruta_excel, read_only=True, data_only=True)
        avisos: list[str] = []
        indice = migrar_productos(session, wb["INVENTARIO"], avisos)
        n_movs = migrar_historial(session, wb["HISTORIAL_"], indice, avisos)
        session.commit()

        n_prod = session.query(Producto).count()
        n_ppc = session.query(Producto).filter(Producto.piezas_por_caja.isnot(None)).count()
        n_lotes = session.query(Lote).count()
        sin_ppc = [p.nombre for p in session.query(Producto)
                   .filter(Producto.piezas_por_caja.is_(None), Producto.unidad == "pieza")]

    with open("migracion_revision.txt", "w", encoding="utf-8") as f:
        f.write("=== AVISOS DE MIGRACIÓN ===\n")
        f.write("\n".join(avisos) or "(sin avisos)")
        f.write(f"\n\n=== PRODUCTOS unidad=pieza SIN piezas_por_caja detectado ({len(sin_ppc)}) ===\n")
        f.write("\n".join(sin_ppc))

    print(f"productos: {n_prod} | lotes: {n_lotes} | movimientos: {n_movs}")
    print(f"piezas_por_caja detectado: {n_ppc} | avisos: {len(avisos)} (ver migracion_revision.txt)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit('Uso: python -m app.migrate "excel original/maybe.xlsm"')
    run(sys.argv[1])
